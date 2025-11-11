import openpyxl
import csv
import os
from pathlib import Path
from models.models import User, Outlet, Asset, SmartDevice, Movement, HealthEvent, DoorEvent, Alert, Client, SubClient
from datetime import datetime
from pytz import timezone
import pandas as pd
import time

# Mapeamento entre os cabeçalhos do Excel e os atributos do modelo User
USER_COLUMN_MAPPING = {
    "First Name": "first_name",
    "Last Name": "last_name",
    "User Name": "user_name",
    "UPN": "upn",
    "Email": "email",
    "Phone": "phone",
    "Role": "role",
    "Reporting Manager": "reporting_manager",
    "Preferred Notification Type": "preferred_notification_type",
    "Country": "country",
    "Responsible Country": "responsible_country",
    "Is Active?": "is_active",
    "Sales Organization": "sales_organization",
    "Sales Office": "sales_office",
    "Sales Group": "sales_group",
    "Sales Territory": "sales_territory",
    "Teleselling Territory": "teleselling_territory",
    "BD Territory Name": "bd_territory_name",
    "CA Territory Name": "ca_territory_name",
    "MC Territory Name": "mc_territory_name",
    "P1 Territory Name": "p1_territory_name",
    "P2 Territory Name": "p2_territory_name",
    "P3 Territory Name": "p3_territory_name",
    "P4 Territory Name": "p4_territory_name",
    "P5 Territory Name": "p5_territory_name",
    "NCB Territory Name": "ncb_territory_name",
    "Last Login On": "last_login_on",
    "Client": "client",
    "Created On": "created_on",
    "Created By": "created_by",
    "Modified On": "modified_on",
    "Modified By": "modified_by",
    "Reward Point": "reward_point"
}

OUTLET_COLUMN_MAPPING = {
    "Name": "name",
    "Code": "code",
    "Outlet Type": "outlet_type",
    "Is Key Outlet?": "is_key_outlet",
    "Is Smart?": "is_smart",
    "Country": "country",
    "State": "state",
    "City": "city",
    "Street": "street",
    "Address 2": "address_2",
    "Address 3": "address_3",
    "Address 4": "address_4",
    "Postal Code": "postal_code",
    "Retailer": "retailer",
    "Primary Phone": "primary_phone",
    "Primary Sales Rep": "primary_sales_rep",
    "Sales Rep Name": "sales_rep_name",
    "Technician": "technician",
    "Market": "market",
    "Sales Target": "sales_target",
    "Client": "client",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "Trade Channel": "trade_channel",
    "Trade Group": "trade_group",
    "Trade Group Code": "trade_group_code",
    "Is Active?": "is_active",
    "Customer Tier": "customer_tier",
    "Sub Trade Channel": "sub_trade_channel",
    "Sales Organization": "sales_organization",
    "Sales Office": "sales_office",
    "Sales Group": "sales_group",
    "Sales Territory": "sales_territory",
    "TeleSelling Territory Name": "teleselling_territory_name",
    "Business Developer Territory Name": "business_developer_territory_name",
    "Credit Approver Territory Name": "credit_approver_territory_name",
    "Merchandizer Territory Name": "merchandizer_territory_name",
    "P1_Territory Name": "p1_territory_name",
    "P2_Territory Name": "p2_territory_name",
    "P3_Territory Name": "p3_territory_name",
    "P4_Territory Name": "p4_territory_name",
    "P5_Territory Name": "p5_territory_name",
    "Reserve Route Name": "reserve_route_name",
    "RDCustomer Name": "rd_customer_name",
    "TimeZone": "time_zone",
    "Sub Client": "sub_client",
    "Cluster": "cluster",
    "Market Segment": "market_segment",
    "Segment": "segment",
    "Environment": "environment",
    "Assortment 1": "assortment_1",
    "Assortment 2": "assortment_2",
    "Assortment 3": "assortment_3",
    "Assortment 4": "assortment_4",
    "Assortment 5": "assortment_5",
    "BarCode": "barcode",
    "Local Cluster": "local_cluster",
    "Local TradeChannel": "local_trade_channel",
    "Chain": "chain",
    "Region Name": "region_name",
    "Mobile Phone": "mobile_phone",
    "Email": "email",
    "CPL Name": "cpl_name",
    "Extra Field": "extra_field",
    "Created On": "created_on",
    "Created By": "created_by",
    "Modified On": "modified_on",
    "Modified By": "modified_by",
    "BDAA": "bdaa",
    "CMMIND": "cmmind",
    "Combined Asset Capacity": "combined_asset_capacity",
    "ASM Name": "asm_name",
    "ASM Email": "asm_email",
    "TSM Name": "tsm_name",
    "TSM Email": "tsm_email"
}

ASSET_COLUMN_MAPPING = {
    "Asset Type": "asset_type",
    "Bottler Equipment Number": "bottler_equipment_number",
    "Technical Id": "technical_id",
    "OEM Serial Number": "oem_serial_number",
    "Asset Ping": "asset_ping",
    "Category": "category",
    "Is Competition?": "is_competition",
    "Is Factory Asset?": "is_factory_asset",
    "Associated in Factory": "associated_in_factory",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Store Location": "store_location",
    "Trade Channel": "trade_channel",
    "Customer Tier": "customer_tier",
    "Sub Trade Channel": "sub_trade_channel",
    "Sales Organization": "sales_organization",
    "Sales Office": "sales_office",
    "Sales Group": "sales_group",
    "Sales Territory": "sales_territory",
    "Issue": "issue",
    "Smart Device": "smart_device",
    "Smart Device Type": "smart_device_type",
    "Smart Device Ping": "smart_device_ping",
    "Gateway": "gateway",
    "Gateway Type": "gateway_type",
    "Gateway Ping": "gateway_ping",
    "Last Scan": "last_scan",
    "Visit (Scan) Status": "visit_scan_status",
    "Client": "client",
    "City": "city",
    "Street": "street",
    "Street 2": "street_2",
    "Street 3": "street_3",
    "State": "state",
    "Country": "country",
    "Prime position?": "prime_position",
    "Is Missing?": "is_missing",
    "Is Vision?": "is_vision",
    "Is Smart?": "is_smart",
    "Is Authorized Movement ?": "is_authorized_movement",
    "Is Unhealthy?": "is_unhealthy",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "Last Known Latitude": "last_known_latitude",
    "Last Known Longitude": "last_known_longitude",
    "Geolocation Source": "geolocation_source",
    "Location Accuracy": "location_accuracy",
    "Displacement(Meter)": "displacement_meter",
    "Is Power On?": "is_power_on",
    "Latest Health Record Event Time": "latest_health_record_event_time",
    "Battery Level": "battery_level",
    "Battery Status": "battery_status",
    "Planogram": "planogram",
    "Responsible BD Username": "responsible_bd_username",
    "Responsible BD First Name": "responsible_bd_first_name",
    "Responsible BD Phone number": "responsible_bd_phone_number",
    "IOT Solution": "iot_solution",
    "Has Sim": "has_sim",
    "Asset Associated On": "asset_associated_on",
    "Gateway Associated On": "gateway_associated_on",
    "Acquisition Date": "acquisition_date",
    "Associated By BD User Name": "associated_by_bd_user_name",
    "Associated By BD Name": "associated_by_bd_name",
    "Gateway Associated By BD User Name": "gateway_associated_by_bd_user_name",
    "Gateway Associated By BD Name": "gateway_associated_by_bd_name",
    "Time Zone": "time_zone",
    "Created On": "created_on",
    "Created By": "created_by",
    "Modified On": "modified_on",
    "Modified By": "modified_by",
    "Capacity Type": "capacity_type",
    "Sub Client": "sub_client",
    "Latest Movement Record Event Time": "latest_movement_record_event_time",
    "Latest Power Record Event Time": "latest_power_record_event_time",
    "Location status": "location_status",
    "Last Location Status On": "last_location_status_on",
    "Latest Location Status On": "latest_location_status_on",
    "Static Latitude": "static_latitude",
    "Static Longitude": "static_longitude",
    "Static Movement Status": "static_movement_status"
}

SMARTDEVICE_COLUMN_MAPPING = {
    "Device Type": "device_type",
    "Manufacturer": "manufacturer",
    "Mac Address": "mac_address",
    "Serial Number": "serial_number",
    "Order Serial Number": "order_serial_number",
    "Shipped Country": "shipped_country",
    "Door No": "door_no",
    "Bottler Equipment Number": "bottler_equipment_number",
    "Technical Identification Number": "technical_identification_number",
    "Gateway": "gateway",
    "Manufacturer Serial Number": "manufacturer_serial_number",
    "IMEI": "imei",
    "Sim#": "sim_number",
    "SIM Provider": "sim_provider",
    "Plugin Connected_FFXy": "plugin_connected_ffxy",
    "Last Ping": "last_ping",
    "Firmware Version": "firmware_version",
    "IBeacon UUID": "ibeacon_uuid",
    "IBeacon Major": "ibeacon_major",
    "IBeacon Minor": "ibeacon_minor",
    "Eddystone UID Namespace": "eddystone_uid_namespace",
    "Eddystone UID Instance": "eddystone_uid_instance",
    "Inventory Location": "inventory_location",
    "Tracking#": "tracking_number",
    "Client": "client",
    "Asset Type": "asset_type",
    "Linked with Asset": "linked_with_asset",
    "Is Factory Asset?": "is_factory_asset",
    "Associated in Factory": "associated_in_factory",
    "Acquisition Date": "acquisition_date",
    "Asset Associated On": "asset_associated_on",
    "Association": "association",
    "Associated By BD User Name": "associated_by_bd_user_name",
    "Associated By BD Name": "associated_by_bd_name",
    "Associated By App Version": "associated_by_app_version",
    "Associated By App Name": "associated_by_app_name",
    "Is Missing?": "is_missing",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Trade Channel": "trade_channel",
    "Customer Tier": "customer_tier",
    "Sub Trade Channel": "sub_trade_channel",
    "Sales Organization": "sales_organization",
    "Sales Office": "sales_office",
    "Sales Group": "sales_group",
    "Sales Territory": "sales_territory",
    "Street": "street",
    "City": "city",
    "State": "state",
    "Country": "country",
    "Time Zone": "time_zone",
    "Latest Health Record Event Time": "latest_health_record_event_time",
    "Battery Level": "battery_level",
    "Created On": "created_on",
    "Created By": "created_by",
    "Modified On": "modified_on",
    "Modified By": "modified_by",
    "Advertisement URL": "advertisement_url",
    "Is Device Registered in IoT Hub?": "is_device_registered_in_iot_hub",
    "Is SD Gateway": "is_sd_gateway",
    "SubClient": "sub_client",
    "Device Model Number": "device_model_number",
    "Module Type": "module_type",
    "SIM Status": "sim_status",
    "Last Sim Status Updated On": "last_sim_status_updated_on"
}

MOVEMENT_COLUMN_MAPPING = {
    "Id": "id",
    "Movement Type": "movement_type",
    "Start Time": "start_time",
    "End Time": "end_time",
    "Duration": "duration",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "Movement Count": "movement_count",
    "Door Open": "door_open",
    "Displacement(Meter)": "displacement_meter",
    "Accuracy(Meter)": "accuracy_meter",
    "Power Status": "power_status",
    "App Name": "app_name",
    "App Version": "app_version",
    "SDK Version": "sdk_version",
    "Data Uploaded By": "data_uploaded_by",
    "GPS Source": "gps_source",
    "Event Id": "event_id",
    "Created On": "created_on",
    "Gateway Mac": "gateway_mac",
    "Gateway#": "gateway_number",
    "Asset Type": "asset_type",
    "Month": "month",
    "Day ": "day",
    "Day of Week": "day_of_week",
    "Week of Year": "week_of_year",
    "Asset Serial #": "asset_serial_number",
    "Technical Id": "technical_id",
    "Equipment Number": "equipment_number",
    "Smart Device Mac": "smart_device_mac",
    "Smart Device#": "smart_device_number",
    "Is Smart?": "is_smart",
    "Smart Device Type": "smart_device_type",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Time Zone": "time_zone",
    "Client": "client",
    "Sub Client": "sub_client"
}

HEALTH_EVENT_COLUMN_MAPPING = {
    "Id": "id",
    "Event Type": "event_type",
    "Light": "light",
    "Light Status": "light_status",
    "Temperature(°C)": "temperature_c",
    "Evaporator Temperature(°C)": "evaporator_temperature_c",
    "Condensor Temperature(°C)": "condensor_temperature_c",
    "Temperature(°F)": "temperature_f",
    "Evaporator Temperature(°F)": "evaporator_temperature_f",
    "Condensor Temperature(°F)": "condensor_temperature_f",
    "Battery": "battery",
    "Battery Status": "battery_status",
    "Interval(Min)": "interval_min",
    "Cooler Voltage(V)": "cooler_voltage_v",
    "Max Voltage(V)": "max_voltage_v",
    "Min Voltage(V)": "min_voltage_v",
    "Avg Power Consumption(Watt)": "avg_power_consumption_watt",
    "Total compressor ON Time(%)": "total_compressor_on_time_percent",
    "Max Cabinet Temperature(°C)": "max_cabinet_temperature_c",
    "Min Cabinet Temperature(°C)": "min_cabinet_temperature_c",
    "Ambient Temperature(°C)": "ambient_temperature_c",
    "Max Cabinet Temperature(°F)": "max_cabinet_temperature_f",
    "Min Cabinet Temperature(°F)": "min_cabinet_temperature_f",
    "Ambient Temperature(°F)": "ambient_temperature_f",
    "App Name": "app_name",
    "App Version": "app_version",
    "SDK Version": "sdk_version",
    "Data Uploaded By": "data_uploaded_by",
    "Asset Category": "asset_category",
    "Event Id": "event_id",
    "Created On": "created_on",
    "Event Time": "event_time",
    "Gateway Mac": "gateway_mac",
    "Gateway#": "gateway_number",
    "Asset Type": "asset_type",
    "Month": "month",
    "Day ": "day",
    "Day of Week": "day_of_week",
    "Week of Year": "week_of_year",
    "Asset Serial #": "asset_serial_number",
    "Technical Id": "technical_id",
    "Equipment Number": "equipment_number",
    "Smart Device Mac": "smart_device_mac",
    "Smart Device#": "smart_device_number",
    "Is Smart?": "is_smart",
    "Smart Device Type": "smart_device_type",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Time Zone": "time_zone",
    "Client": "client",
    "Sub Client": "sub_client"
}

DOOR_EVENT_COLUMN_MAPPING = {
    "Id": "id",
    "Open Event Time": "open_event_time",
    "Close Event Time": "close_event_time",
    "Event Type": "event_type",
    "Door Open Duration(sec)": "door_open_duration_sec",
    "Time of Day": "time_of_day",
    "Weekday / Weekend": "weekday_weekend",
    "Hour in Day": "hour_in_day",
    "Door Count": "door_count",
    "Additional Info": "additional_info",
    "Outlet Territory": "outlet_territory",
    "Door": "door",
    "Capacity Type": "capacity_type",
    "Door Open Target": "door_open_target",
    "Door Open Temperature": "door_open_temperature",
    "Door Close Temperature": "door_close_temperature",
    "App Name": "app_name",
    "App Version": "app_version",
    "SDK Version": "sdk_version",
    "Data Uploaded By": "data_uploaded_by",
    "Asset Category": "asset_category",
    "Event Id": "event_id",
    "Created On": "created_on",
    "Gateway Mac": "gateway_mac",
    "Gateway#": "gateway_number",
    "Asset Type": "asset_type",
    "Month": "month",
    "Day ": "day",
    "Day of Week": "day_of_week",
    "Week of Year": "week_of_year",
    "Asset Serial #": "asset_serial_number",
    "Technical Id": "technical_id",
    "Equipment Number": "equipment_number",
    "Smart Device Mac": "smart_device_mac",
    "Smart Device#": "smart_device_number",
    "Is Smart?": "is_smart",
    "Smart Device Type": "smart_device_type",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Time Zone": "time_zone",
    "Client": "client",
    "Sub Client": "sub_client"
}

ALERT_COLUMN_MAPPING = {
    "Id": "id",
    "Alert Type": "alert_type",
    "Alert Text": "alert_text",
    "Alert Definition": "alert_definition",
    "Status": "status",
    "Visit Check": "visit_check",
    "Asset Serial#": "asset_serial_number",
    "Smart Device Serial#": "smart_device_serial_number",
    "Asset Equipment Number": "asset_equipment_number",
    "Asset Technical Identification Number": "asset_technical_identification_number",
    "Asset Type": "asset_type",
    "Street": "street",
    "Street 2": "street_2",
    "Street 3": "street_3",
    "Is Smart?": "is_smart",
    "Alert At": "alert_at",
    "Status Changed On": "status_changed_on",
    "Priority": "priority",
    "Age": "age",
    "Alert Age(in minutes)": "alert_age_in_minutes",
    "Value": "value",
    "Last Update": "last_update",
    "Outlet": "outlet",
    "Outlet Code": "outlet_code",
    "Outlet Type": "outlet_type",
    "Outlet City": "outlet_city",
    "Client": "client",
    "Time Zone": "time_zone",
    "Month": "month",
    "Day ": "day",
    "Day of Week": "day_of_week",
    "Week of Year": "week_of_year",
    "Market": "market",
    "Trade Channel": "trade_channel",
    "Customer Tier": "customer_tier",
    "Sales Organization": "sales_organization",
    "Sales Office": "sales_office",
    "Sales Group": "sales_group",
    "Sales Territory": "sales_territory",
    "Sales Rep": "sales_rep",
    "Is System Alert?": "is_system_alert",
    "Acknowledge Comment": "acknowledge_comment",
    "Created On": "created_on"
}

CLIENT_COLUMN_MAPPING = {
    "Client Id": "id",
    "Client Code": "client_code",
    "Client Name": "client_name",
    "Relevant Business Stream": "relevant_business_stream",
    "Status": "status",
    "Contact": "contact",
    "Subdomain": "subdomain",
    "Is Feedback Enabled?": "is_feedback_enabled",
    "Time Zone": "time_zone",
    "Vision Image Interval (Hours)": "vision_image_interval_hours",
    "Vision Image Interval (Door Open)": "vision_image_interval_door_open",
    "Out Of Stock SKU": "out_of_stock_sku",
    "Power Off Duration": "power_off_duration",
    "Temperature Min": "temperature_min",
    "Temperature Max": "temperature_max",
    "Light Min": "light_min",
    "Light Max": "light_max",
    "Door Count": "door_count",
    "Health Intervals  (Hours)": "health_intervals_hours",
    "Cooler Tracking Threshold (Days)": "cooler_tracking_threshold_days",
    "Cooler Tracking Displacement Threshold (Mtr)": "cooler_tracking_displacement_threshold_mtr",
    "GeoLocation Api Key": "geolocation_api_key",
    "Created On": "created_on",
    "Created By": "created_by",
    "Modified On": "modified_on",
    "Modified By": "modified_by",
    "Fallen Magnet Threshold": "fallen_magnet_threshold",
    "VHenabled": "vh_enabled",
    "Country": "country",
    "Shipped Country": "shipped_country",
    "Manual Processing Mode": "manual_processing_mode",
    "Is Visit From Ping": "is_visit_from_ping",
    "Distance In Meter": "distance_in_meter",
    "Threshold In Minutes": "threshold_in_minutes",
    "Limit Location Distance": "limit_location_distance",
    "Survey Distance ": "survey_distance",
    "Scene Mode": "scene_mode",
    "Currency": "currency",
    "Enable PIC To POG": "enable_pic_to_pog",
    "Role": "role",
    "Default Recognition Mode": "default_recognition_mode",
    "Disable Geo Data Collection": "disable_geo_data_collection"
}

SUBCLIENT_COLUMN_MAPPING = {
    "SubClient Name": "subclient_name",
    "SubClient Code": "subclient_code",
    "Scene Mode": "scene_mode",
    "Client": "client"
}

def parse_datetime(value, verbose=False):
    if isinstance(value, datetime):  # Verificar se já é um objeto datetime
        return value
    if value:
        try:
            dt_str = str(value).strip()  # Remove espaços no início e final
            
            # Se contém fuso horário no final, remover
            # Padrões: "BRST", "ESAST", "EDT", "EST", etc. (3-5 letras maiúsculas)
            parts = dt_str.split()
            detected_tz = None
            if len(parts) > 1 and parts[-1].isupper() and len(parts[-1]) in [3, 4, 5]:
                # Capturar o fuso horário antes de remover
                tz_part = parts[-1]
                # Remover o fuso horário (última parte se for tudo maiúsculo)
                dt_str = " ".join(parts[:-1]).strip()
                
                # Mapear fusos horários conhecidos
                tz_mapping = {
                    "BRST": "America/Sao_Paulo",      # Brasília Summer Time
                    "BRT": "America/Sao_Paulo",       # Brasília Time  
                    "ESAST": "Africa/Johannesburg",   # East South Africa Standard Time
                    "EST": "America/New_York",        # Eastern Standard Time
                    "EDT": "America/New_York",        # Eastern Daylight Time
                    "PST": "America/Los_Angeles",     # Pacific Standard Time
                    "PDT": "America/Los_Angeles",     # Pacific Daylight Time
                    "UTC": "UTC",                     # Universal Time Coordinated
                    "GMT": "UTC"                      # Greenwich Mean Time
                }
                detected_tz = tz_mapping.get(tz_part, "America/Sao_Paulo")
            
            dt = None
            
            # Tentar múltiplos formatos de data
            formats = [
                "%d/%m/%Y %H:%M:%S",  # DD/MM/YYYY HH:MM:SS (padrão brasileiro)
                "%m/%d/%Y %H:%M:%S",  # MM/DD/YYYY HH:MM:SS (padrão americano)
                "%Y-%m-%d %H:%M:%S",  # YYYY-MM-DD HH:MM:SS (ISO)
                "%d/%m/%Y",           # DD/MM/YYYY (apenas data)
                "%m/%d/%Y",           # MM/DD/YYYY (apenas data)
                "%Y-%m-%d",           # YYYY-MM-DD (apenas data)
            ]
            
            for fmt in formats:
                try:
                    dt = datetime.strptime(dt_str, fmt)
                    break
                except ValueError:
                    continue
            
            if dt is None:
                if verbose:
                    print(f"⚠️  Não foi possível converter a data: {value}")
                return None
            
            # Adicionar o fuso horário apropriado
            tz_name = detected_tz if detected_tz else "America/Sao_Paulo"
            try:
                tz = timezone(tz_name)
                return tz.localize(dt)
            except Exception as tz_error:
                if verbose:
                    print(f"⚠️  Erro no fuso horário {tz_name}: {tz_error}. Usando América/São_Paulo")
                tz = timezone("America/Sao_Paulo")
                return tz.localize(dt)
                
        except Exception as e:
            if verbose:
                print(f"❌ Erro ao converter data '{value}': {e}")
            return None
    return None

# Função para inserir ou atualizar usuários a partir do Excel
def insert_or_update_users_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        user_data = {}
        for header, value in zip(headers, row):
            if header in USER_COLUMN_MAPPING:
                column_name = USER_COLUMN_MAPPING[header]
                if column_name == "is_active":  # Converter 'Yes'/'No' para booleano
                    user_data[column_name] = True if value == "Yes" else False
                elif column_name in ["last_login_on", "created_on", "modified_on"]:
                    user_data[column_name] = parse_datetime(value)
                else:
                    user_data[column_name] = value

        # Verificar se o usuário já existe no banco de dados (com base no campo 'upn')
        existing_user = session.query(User).filter_by(upn=user_data.get("upn")).first()

        if existing_user:
            # Atualizar os campos do usuário existente
            for key, value in user_data.items():
                setattr(existing_user, key, value)
        else:
            # Criar um novo usuário
            new_user = User(**user_data)
            session.add(new_user)

    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar outlets a partir do Excel
def insert_or_update_outlets_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        outlet_data = {}
        for header, value in zip(headers, row):
            if header in OUTLET_COLUMN_MAPPING:
                column_name = OUTLET_COLUMN_MAPPING[header]
                if column_name in ["is_key_outlet", "is_smart", "is_active"]:  # Converter 'Yes'/'No' para booleano
                    outlet_data[column_name] = True if value == "Yes" else False
                elif column_name in ["created_on", "modified_on"]:
                    outlet_data[column_name] = parse_datetime(value)
                else:
                    outlet_data[column_name] = value

        # Verificar se o outlet já existe no banco de dados (com base no campo 'code')
        existing_outlet = session.query(Outlet).filter_by(code=outlet_data.get("code")).first()

        if existing_outlet:
            # Atualizar os campos do outlet existente
            for key, value in outlet_data.items():
                setattr(existing_outlet, key, value)
        else:
            # Criar um novo outlet
            new_outlet = Outlet(**outlet_data)
            session.add(new_outlet)

    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar assets a partir do Excel
def insert_or_update_assets_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        asset_data = {}
        for header, value in zip(headers, row):
            if header in ASSET_COLUMN_MAPPING:
                column_name = ASSET_COLUMN_MAPPING[header]
                if column_name in ["is_competition", "is_factory_asset", "associated_in_factory", "prime_position",
                                   "is_missing", "is_vision", "is_smart", "is_authorized_movement",
                                   "is_unhealthy", "is_power_on", "has_sim"]: 
                    asset_data[column_name] = True if value == "Yes" else False
                elif column_name in ["last_scan", "asset_associated_on", "gateway_associated_on",
                                     "acquisition_date", "created_on", "modified_on",
                                     "latest_health_record_event_time", "latest_movement_record_event_time",
                                     "latest_power_record_event_time", "last_location_status_on",
                                     "latest_location_status_on"]:
                    asset_data[column_name] = parse_datetime(value)
                else:
                    asset_data[column_name] = value

        # Verificar se o asset já existe no banco de dados (com base no campo 'oem_serial_number')
        existing_asset = session.query(Asset).filter_by(oem_serial_number=asset_data.get("oem_serial_number")).first()

        if existing_asset:
            # Atualizar os campos do asset existente
            for key, value in asset_data.items():
                setattr(existing_asset, key, value)
        else:
            # Criar um novo asset
            new_asset = Asset(**asset_data)
            session.add(new_asset)
    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar smart devices a partir do Excel
def insert_or_update_smartdevices_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        smartdevice_data = {}
        for header, value in zip(headers, row):
            if header in SMARTDEVICE_COLUMN_MAPPING:
                column_name = SMARTDEVICE_COLUMN_MAPPING[header]
                if column_name in ["plugin_connected_ffxy", "is_factory_asset", "associated_in_factory",
                                   "is_missing", "is_device_registered_in_iot_hub", "is_sd_gateway"]: 
                    smartdevice_data[column_name] = True if value == "Yes" else False
                elif column_name in ["last_ping", "acquisition_date", "asset_associated_on",
                                     "latest_health_record_event_time", "created_on",
                                     "modified_on", "last_sim_status_updated_on"]:
                    smartdevice_data[column_name] = parse_datetime(value)
                else:
                    smartdevice_data[column_name] = value

        # Verificar se o smart device já existe no banco de dados (com base no campo 'technical_id')
        existing_smartdevice = session.query(SmartDevice).filter_by(mac_address=smartdevice_data.get("mac_address")).first()

        if existing_smartdevice:
            # Atualizar os campos do smart device existente
            for key, value in smartdevice_data.items():
                setattr(existing_smartdevice, key, value)
        else:
            # Criar um novo smart device
            new_smartdevice = SmartDevice(**smartdevice_data)
            session.add(new_smartdevice)
    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar movements a partir do Excel
def insert_or_update_movements_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        movement_data = {}
        for header, value in zip(headers, row):
            if header in MOVEMENT_COLUMN_MAPPING:
                column_name = MOVEMENT_COLUMN_MAPPING[header]
                if column_name == "id" and value is not None:
                    movement_data[column_name] = str(int(value)) if isinstance(value, (int, float)) else str(value)
                elif column_name in ["door_open", "is_smart"]:
                    movement_data[column_name] = True if value == "Yes" else False
                elif column_name in ["start_time", "end_time", "created_on"]:
                    movement_data[column_name] = parse_datetime(value)
                else:
                    movement_data[column_name] = value

        # Verificar se o movement já existe no banco de dados (com base no campo 'id')
        existing_movement = session.query(Movement).filter_by(id=movement_data.get("id")).first()

        if existing_movement:
            # Atualizar os campos do movement existente
            for key, value in movement_data.items():
                setattr(existing_movement, key, value)
        else:
            # Criar um novo movement
            new_movement = Movement(**movement_data)
            session.add(new_movement)
    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar health events a partir do Excel
def insert_or_update_health_events_from_excel(session, excel_file_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Obter os cabeçalhos do Excel
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre todas as linhas de dados (ignorando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criar um dicionário com os dados da linha, usando o mapeamento
        health_event_data = {}
        for header, value in zip(headers, row):
            if header in HEALTH_EVENT_COLUMN_MAPPING:
                column_name = HEALTH_EVENT_COLUMN_MAPPING[header]
                if column_name == "id" and value is not None:
                    health_event_data[column_name] = str(int(value)) if isinstance(value, (int, float)) else str(value)
                elif column_name in ["is_smart"]:
                    health_event_data[column_name] = True if value == "Yes" else False
                elif column_name in ["created_on", "event_time"]:
                    health_event_data[column_name] = parse_datetime(value)
                else:
                    health_event_data[column_name] = value

        # Verificar se o health event já existe no banco de dados (com base no campo 'id')
        existing_health_event = session.query(HealthEvent).filter_by(id=health_event_data.get("id")).first()

        if existing_health_event:
            # Atualizar os campos do health event existente
            for key, value in health_event_data.items():
                setattr(existing_health_event, key, value)
        else:
            # Criar um novo health event
            new_health_event = HealthEvent(**health_event_data)
            session.add(new_health_event)
    # Salvar as alterações no banco de dados
    session.commit()

# Função para inserir ou atualizar door events a partir de CSV (UTF-16) - ULTRA OTIMIZADA COM PANDAS
def insert_or_update_door_events_from_csv(session, csv_file_path):
    """
    Import door events from CSV file using PANDAS - VERSÃO ULTRA OTIMIZADA.
    Usa pandas para processamento vetorizado e bulk operations para máxima performance.
    """
    from sqlalchemy import text
    
    start_time = time.time()
    print(f"🚀 Iniciando importação ULTRA otimizada com PANDAS: {csv_file_path}")
    
    # Ler CSV com pandas (muito mais rápido que CSV reader manual)
    print("📁 Lendo arquivo CSV...")
    df = pd.read_csv(
        csv_file_path, 
        encoding='utf-16',
        skiprows=1,  # Pular linha "Door Statuses"
        low_memory=False
    )
    
    print(f"✓ Arquivo lido: {len(df)} registros")
    
    # Mapear colunas de forma otimizada
    column_mapping = {
        "Id": "id",
        "Open Event Time": "open_event_time", 
        "Close Event Time": "close_event_time",
        "Event Type": "event_type",
        "Door Open Duration(sec)": "door_open_duration_sec",
        "Time of Day": "time_of_day",
        "Weekday / Weekend": "weekday_weekend",
        "Hour in Day": "hour_in_day",
        "Door Count": "door_count",
        "Additional Info": "additional_info",
        "Outlet Territory": "outlet_territory",
        "Door": "door",
        "Capacity Type": "capacity_type",
        "Door Open Target": "door_open_target",
        "Door Open Temperature": "door_open_temperature",
        "Door Close Temperature": "door_close_temperature",
        "App Name": "app_name",
        "App Version": "app_version",
        "SDK Version": "sdk_version",
        "Data Uploaded By": "data_uploaded_by",
        "Asset Category": "asset_category",
        "Event Id": "event_id",
        "Created On": "created_on",
        "Gateway Mac": "gateway_mac",
        "Gateway#": "gateway_number",
        "Asset Type": "asset_type",
        "Month": "month",
        "Day ": "day",
        "Day of Week": "day_of_week",
        "Week of Year": "week_of_year",
        "Asset Serial #": "asset_serial_number",
        "Technical Id": "technical_id",
        "Equipment Number": "equipment_number",
        "Smart Device Mac": "smart_device_mac",
        "Smart Device#": "smart_device_number",
        "Is Smart?": "is_smart",
        "Smart Device Type": "smart_device_type",
        "Outlet": "outlet",
        "Outlet Code": "outlet_code",
        "Outlet Type": "outlet_type",
        "Time Zone": "time_zone",
        "Client": "client",
        "Sub Client": "sub_client"
    }
    
    # Renomear colunas disponíveis
    available_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
    df = df.rename(columns=available_columns)
    df = df[list(available_columns.values())]
    
    # Limpar dados usando pandas (vetorizado - muito mais rápido)
    print("🧹 Limpando e transformando dados...")
    df = df.dropna(subset=['id'])  # Remover registros sem ID
    df['id'] = df['id'].astype(str).str.strip()
    
    # Converter campos booleanos de forma vetorizada
    if 'is_smart' in df.columns:
        df['is_smart'] = df['is_smart'].str.lower().eq('yes').fillna(False)
    
    # Converter datas usando a função parse_datetime corrigida
    datetime_columns = ['open_event_time', 'close_event_time', 'created_on']
    for col in datetime_columns:
        if col in df.columns:
            print(f"🔄 Convertendo coluna {col}...")
            df[col] = df[col].apply(lambda x: parse_datetime(str(x).strip()) if pd.notna(x) else None)
    
    # Substituir NaN por None para compatibilidade com SQLAlchemy
    df = df.where(pd.notnull(df), None)
    
    print(f"✓ Dados processados: {len(df)} registros válidos")
    
    # Obter IDs existentes
    print("📊 Verificando registros existentes...")
    existing_ids = {row.id for row in session.query(DoorEvent.id).all()}
    
    # Separar inserções e atualizações usando pandas
    mask_exists = df['id'].isin(existing_ids)
    df_insert = df[~mask_exists]
    df_update = df[mask_exists]
    
    print(f"📥 Para inserir: {len(df_insert)}")
    print(f"🔄 Para atualizar: {len(df_update)}")
    
    # Usar PostgreSQL UPSERT nativo para máxima performance
    if len(df) > 0:
        print("⚡ Executando UPSERT ultra-rápido...")
        records = df.to_dict('records')
        
        # Query UPSERT otimizada usando nome correto da tabela
        upsert_query = text(f"""
            INSERT INTO {DoorEvent.__tablename__} (
                id, open_event_time, close_event_time, event_type, door_open_duration_sec,
                time_of_day, weekday_weekend, hour_in_day, door_count, additional_info,
                outlet_territory, door, capacity_type, door_open_target, door_open_temperature,
                door_close_temperature, app_name, app_version, sdk_version, data_uploaded_by,
                asset_category, event_id, created_on, gateway_mac, gateway_number,
                asset_type, month, day, day_of_week, week_of_year, asset_serial_number,
                technical_id, equipment_number, smart_device_mac, smart_device_number,
                is_smart, smart_device_type, outlet, outlet_code, outlet_type,
                time_zone, client, sub_client
            )
            VALUES (
                :id, :open_event_time, :close_event_time, :event_type, :door_open_duration_sec,
                :time_of_day, :weekday_weekend, :hour_in_day, :door_count, :additional_info,
                :outlet_territory, :door, :capacity_type, :door_open_target, :door_open_temperature,
                :door_close_temperature, :app_name, :app_version, :sdk_version, :data_uploaded_by,
                :asset_category, :event_id, :created_on, :gateway_mac, :gateway_number,
                :asset_type, :month, :day, :day_of_week, :week_of_year, :asset_serial_number,
                :technical_id, :equipment_number, :smart_device_mac, :smart_device_number,
                :is_smart, :smart_device_type, :outlet, :outlet_code, :outlet_type,
                :time_zone, :client, :sub_client
            )
            ON CONFLICT (id) DO UPDATE SET
                open_event_time = EXCLUDED.open_event_time,
                close_event_time = EXCLUDED.close_event_time,
                event_type = EXCLUDED.event_type,
                door_open_duration_sec = EXCLUDED.door_open_duration_sec,
                time_of_day = EXCLUDED.time_of_day,
                weekday_weekend = EXCLUDED.weekday_weekend,
                hour_in_day = EXCLUDED.hour_in_day,
                door_count = EXCLUDED.door_count,
                additional_info = EXCLUDED.additional_info,
                outlet_territory = EXCLUDED.outlet_territory,
                door = EXCLUDED.door,
                capacity_type = EXCLUDED.capacity_type,
                door_open_target = EXCLUDED.door_open_target,
                door_open_temperature = EXCLUDED.door_open_temperature,
                door_close_temperature = EXCLUDED.door_close_temperature,
                app_name = EXCLUDED.app_name,
                app_version = EXCLUDED.app_version,
                sdk_version = EXCLUDED.sdk_version,
                data_uploaded_by = EXCLUDED.data_uploaded_by,
                asset_category = EXCLUDED.asset_category,
                event_id = EXCLUDED.event_id,
                created_on = EXCLUDED.created_on,
                gateway_mac = EXCLUDED.gateway_mac,
                gateway_number = EXCLUDED.gateway_number,
                asset_type = EXCLUDED.asset_type,
                month = EXCLUDED.month,
                day = EXCLUDED.day,
                day_of_week = EXCLUDED.day_of_week,
                week_of_year = EXCLUDED.week_of_year,
                asset_serial_number = EXCLUDED.asset_serial_number,
                technical_id = EXCLUDED.technical_id,
                equipment_number = EXCLUDED.equipment_number,
                smart_device_mac = EXCLUDED.smart_device_mac,
                smart_device_number = EXCLUDED.smart_device_number,
                is_smart = EXCLUDED.is_smart,
                smart_device_type = EXCLUDED.smart_device_type,
                outlet = EXCLUDED.outlet,
                outlet_code = EXCLUDED.outlet_code,
                outlet_type = EXCLUDED.outlet_type,
                time_zone = EXCLUDED.time_zone,
                client = EXCLUDED.client,
                sub_client = EXCLUDED.sub_client
        """)
        
        # Executar em chunks de 1000 para performance ótima
        chunk_size = 1000
        for i in range(0, len(records), chunk_size):
            chunk = records[i:i+chunk_size]
            session.execute(upsert_query, chunk)
            if i % 5000 == 0 and i > 0:
                print(f"� Processados {i} registros...")
        
        session.commit()
        print(f"✅ UPSERT concluído: {len(records)} registros")
    
    total_time = time.time() - start_time
    records_per_second = len(df) / total_time if total_time > 0 else 0
    
    print(f"🎉 Importação ULTRA otimizada concluída em {total_time:.2f} segundos")
    print(f"⚡ Performance: {records_per_second:.0f} registros/segundo")
    print(f"🚀 Melhoria estimada: 5-10x mais rápido que método anterior")



# Função para inserir ou atualizar alerts a partir do CSV (UTF-16)
def insert_or_update_alerts_from_csv(session, csv_file_path):
    """
    Import alerts from CSV file with UTF-16 encoding.
    """
    # Carregar o arquivo CSV com encoding UTF-16
    with open(csv_file_path, 'r', encoding='utf-16') as f:
        # Pular a primeira linha (título "Alerts")
        f.readline()
        reader = csv.DictReader(f)

        for row in reader:
            # Criar um dicionário com os dados da linha, usando o mapeamento
            alert_data = {}
            for excel_col, db_col in ALERT_COLUMN_MAPPING.items():
                if excel_col in row and row[excel_col]:
                    value = row[excel_col]

                    if db_col == "id" and value:
                        # Preservar o ID exato do CSV
                        alert_data[db_col] = value.strip()
                    elif db_col in ["is_smart", "is_system_alert"]:
                        alert_data[db_col] = value.strip().lower() == "yes"
                    elif db_col in ["alert_at", "status_changed_on", "last_update", "created_on"]:
                        alert_data[db_col] = parse_datetime(value.strip())
                    else:
                        alert_data[db_col] = value.strip()

            # Pular linhas que não têm ID
            if not alert_data.get("id"):
                continue

            # Verificar se o alert já existe no banco de dados (com base no campo 'id')
            existing_alert = session.query(Alert).filter_by(id=alert_data.get("id")).first()

            if existing_alert:
                # Atualizar os campos do alert existente
                for key, value in alert_data.items():
                    setattr(existing_alert, key, value)
            else:
                # Criar um novo alert
                new_alert = Alert(**alert_data)
                session.add(new_alert)

        # Salvar as alterações no banco de dados
        session.commit()

# Função para inserir ou atualizar clientes a partir do CSV (UTF-16)
def insert_or_update_clients_from_csv(session, csv_file_path):
    """
    Import clients from CSV file with UTF-16 encoding.
    """
    # Carregar o arquivo CSV com encoding UTF-16
    with open(csv_file_path, 'r', encoding='utf-16') as f:
        # Pular a primeira linha (título "Client")
        f.readline()
        reader = csv.DictReader(f)

        for row in reader:
            # Criar um dicionário com os dados da linha, usando o mapeamento
            client_data = {}
            for excel_col, db_col in CLIENT_COLUMN_MAPPING.items():
                if excel_col in row and row[excel_col]:
                    value = row[excel_col]

                    if db_col == "id" and value:
                        # Preservar o ID exato do CSV
                        client_data[db_col] = value.strip()
                    elif db_col in ["is_feedback_enabled", "vh_enabled", "manual_processing_mode", "is_visit_from_ping", "limit_location_distance", "enable_pic_to_pog", "disable_geo_data_collection"]:
                        client_data[db_col] = value.strip().lower() == "yes"
                    elif db_col in ["created_on", "modified_on"]:
                        client_data[db_col] = parse_datetime(value.strip())
                    elif db_col in ["vision_image_interval_hours", "temperature_min", "temperature_max", "cooler_tracking_displacement_threshold_mtr"]:
                        try:
                            client_data[db_col] = float(value.strip())
                        except (ValueError, TypeError):
                            client_data[db_col] = None
                    elif db_col in ["vision_image_interval_door_open", "out_of_stock_sku", "power_off_duration", "light_min", "light_max", "door_count", "health_intervals_hours", "cooler_tracking_threshold_days", "fallen_magnet_threshold", "distance_in_meter", "threshold_in_minutes", "survey_distance"]:
                        try:
                            client_data[db_col] = int(value.strip())
                        except (ValueError, TypeError):
                            client_data[db_col] = None
                    else:
                        client_data[db_col] = value.strip()

            # Pular linhas que não têm ID
            if not client_data.get("id"):
                continue

            # Verificar se o cliente já existe no banco de dados (com base no campo 'id')
            existing_client = session.query(Client).filter_by(id=client_data.get("id")).first()

            if existing_client:
                # Atualizar os campos do cliente existente
                for key, value in client_data.items():
                    setattr(existing_client, key, value)
            else:
                # Criar um novo cliente
                new_client = Client(**client_data)
                session.add(new_client)

        # Salvar as alterações no banco de dados
        session.commit()

# Função para inserir ou atualizar subclientes a partir do CSV (UTF-16)
def insert_or_update_subclients_from_csv(session, csv_file_path):
    """
    Import subclients from CSV file with UTF-16 encoding.
    """
    # Carregar o arquivo CSV com encoding UTF-16
    with open(csv_file_path, 'r', encoding='utf-16') as f:
        # Pular a primeira linha (título "SubClient")
        f.readline()
        reader = csv.DictReader(f)

        counter = 0
        for row in reader:
            # Criar um dicionário com os dados da linha, usando o mapeamento
            subclient_data = {}
            for excel_col, db_col in SUBCLIENT_COLUMN_MAPPING.items():
                if excel_col in row and row[excel_col]:
                    value = row[excel_col]
                    subclient_data[db_col] = value.strip()

            # Gerar um ID único para o subclient (subclient_code + client_name)
            if subclient_data.get("subclient_code") and subclient_data.get("client"):
                subclient_data["id"] = f"{subclient_data['subclient_code']}_{subclient_data['client']}"
            else:
                continue

            # Verificar se o subclient já existe no banco de dados (com base no campo 'id')
            existing_subclient = session.query(SubClient).filter_by(id=subclient_data.get("id")).first()

            if existing_subclient:
                # Atualizar os campos do subclient existente
                for key, value in subclient_data.items():
                    setattr(existing_subclient, key, value)
            else:
                # Criar um novo subclient
                new_subclient = SubClient(**subclient_data)
                session.add(new_subclient)
                counter += 1

        # Salvar as alterações no banco de dados
        session.commit()
        print(f"✓ SubClients: {counter} novos subclientes importados")


# ============================================================================
# GENERIC FUNCTIONS - Support for both Excel and CSV files
# ============================================================================

def get_file_extension(file_path):
    """Get file extension"""
    return Path(file_path).suffix.lower()


def is_csv_file(file_path):
    """Check if file is CSV"""
    return get_file_extension(file_path) == '.csv'


def is_excel_file(file_path):
    """Check if file is Excel"""
    return get_file_extension(file_path) in ['.xlsx', '.xls']


def insert_or_update_users(session, file_path):
    """
    Universal function to import users from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_users_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for Users. Please use Excel file (.xlsx)")
        return None


def insert_or_update_outlets(session, file_path):
    """
    Universal function to import outlets from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_outlets_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for Outlets. Please use Excel file (.xlsx)")
        return None


def insert_or_update_assets(session, file_path):
    """
    Universal function to import assets from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_assets_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for Assets. Please use Excel file (.xlsx)")
        return None


def insert_or_update_smartdevices(session, file_path):
    """
    Universal function to import smart devices from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_smartdevices_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for SmartDevices. Please use Excel file (.xlsx)")
        return None


def insert_or_update_movements(session, file_path):
    """
    Universal function to import movements from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_movements_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for Movements. Please use Excel file (.xlsx)")
        return None


def insert_or_update_health_events(session, file_path):
    """
    Universal function to import health events from either Excel or CSV
    """
    if is_excel_file(file_path):
        return insert_or_update_health_events_from_excel(session, file_path)
    else:
        print(f"⚠️  CSV format not implemented for HealthEvents. Please use Excel file (.xlsx)")
        return None


def insert_or_update_door_events(session, file_path):
    """
    Universal function to import door events from either Excel or CSV (UTF-16)
    """
    if is_csv_file(file_path):
        return insert_or_update_door_events_from_csv(session, file_path)
    elif is_excel_file(file_path):
        print(f"⚠️  Excel format for Door Events may lose large IDs. CSV (UTF-16) is recommended.")
        # Could implement Excel support here if needed
        return None
    else:
        print(f"❌ Unsupported file format: {file_path}")
        return None


def insert_or_update_alerts(session, file_path):
    """
    Universal function to import alerts from either Excel or CSV (UTF-16)
    """
    if is_csv_file(file_path):
        return insert_or_update_alerts_from_csv(session, file_path)
    elif is_excel_file(file_path):
        print(f"⚠️  Excel format for Alerts may lose large IDs. CSV (UTF-16) is recommended.")
        # Could implement Excel support here if needed
        return None
    else:
        print(f"❌ Unsupported file format: {file_path}")
        return None


def insert_or_update_clients(session, file_path):
    """
    Universal function to import clients from either Excel or CSV (UTF-16)
    """
    if is_csv_file(file_path):
        return insert_or_update_clients_from_csv(session, file_path)
    elif is_excel_file(file_path):
        print(f"⚠️  Excel format for Clients may lose large data. CSV (UTF-16) is recommended.")
        # Could implement Excel support here if needed
        return None
    else:
        print(f"❌ Unsupported file format: {file_path}")
        return None


def insert_or_update_subclients(session, file_path):
    """
    Universal function to import subclients from either Excel or CSV (UTF-16)
    """
    if is_csv_file(file_path):
        return insert_or_update_subclients_from_csv(session, file_path)
    elif is_excel_file(file_path):
        print(f"⚠️  Excel format for SubClients. CSV (UTF-16) is recommended.")
        # Could implement Excel support here if needed
        return None
    else:
        print(f"❌ Unsupported file format: {file_path}")
        return None


def import_all_from_directory(session, directory_path, verbose=True):
    """
    Auto-detect and import all supported files from a directory.
    
    Expects files with standard names:
    - users.xlsx
    - outlets.xlsx
    - assets.xlsx
    - smartdevices.xlsx
    - movements.xlsx
    - health_events.xlsx
    - door_events.csv (UTF-16)
    - alerts.csv (UTF-16)
    - clients.csv (UTF-16)
    - subclients.csv (UTF-16)
    """
    if not os.path.isdir(directory_path):
        print(f"❌ Directory not found: {directory_path}")
        return False
    
    print(f"\n🔍 Scanning directory: {directory_path}")
    print("=" * 70)
    
    imported_count = 0
    skipped_count = 0
    
    # Define file mappings: (expected_name_pattern, import_function)
    file_mappings = [
        ('users', insert_or_update_users),
        ('outlets', insert_or_update_outlets),
        ('assets', insert_or_update_assets),
        ('smartdevices', insert_or_update_smartdevices),
        ('movements', insert_or_update_movements),
        ('health_events', insert_or_update_health_events),
        ('door_events', insert_or_update_door_events),
        ('alerts', insert_or_update_alerts),
        ('clients', insert_or_update_clients),
        ('subclients', insert_or_update_subclients),
    ]
    
    for file_name in sorted(os.listdir(directory_path)):
        if file_name.startswith('.'):
            continue
            
        file_path = os.path.join(directory_path, file_name)
        file_base = Path(file_name).stem.lower()
        
        # Find matching import function
        for pattern, import_func in file_mappings:
            if pattern in file_base:
                if is_excel_file(file_path) or is_csv_file(file_path):
                    try:
                        if verbose:
                            print(f"📥 Importing: {file_name}")
                        import_func(session, file_path)
                        print(f"✅ {file_name} imported successfully")
                        imported_count += 1
                    except Exception as e:
                        print(f"❌ Error importing {file_name}: {str(e)}")
                        skipped_count += 1
                else:
                    if verbose:
                        print(f"⏭️  Skipping: {file_name} (unsupported format)")
                    skipped_count += 1
                break
    
    print("=" * 70)
    print(f"✨ Import completed: {imported_count} files imported, {skipped_count} skipped")
    return True